"""
Assumes json file of js-array-object with dict like [{},{}]
All elements should atleast contain the same keys as the 0th!
Fields or Keys of the dictionary should be provided in an array
Fields and save can be changed for reusing functions!
"""

import json
from openpyxl import Workbook

wb = Workbook()

# FIELDS and FILE_LIST in GSoCArchive!
# switchFIELDS = list(filter(lambda x: "16" not in x, FIELDS)) for GSoC16 bug!


def populator(file_list, fields):
    global wb
    for file_name in file_list:
        with open(file_name, 'r+') as working_file:
            jsonfl = json.load(working_file)
        sheet = wb.create_sheet(title=file_name[:-5])    # '.json'
        filler(jsonfl, sheet, fields)
        print(fields)


def filler(jsonfl, sheet, fields):
    global wb
    for colIn, header in enumerate(fields, start=1):
        sheet.cell(row=1, column=colIn, value=header)
    for rowIn, organ in enumerate(jsonfl, start=2):
        for colIn, header in enumerate(fields, start=1):
            try:
                tag = organ[header]
            except KeyError:
                tag = "-"
            sheet.cell(row=rowIn, column=colIn, value=str(tag))


# Save part has been moved to GSoCArchive
